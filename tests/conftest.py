"""
tests/conftest.py
-------------------
Fixtures compartilhadas. Importante: os testes usam dados SINTÉTICOS
(nome/CPF fictícios), nunca o extrato real do usuário — o formato é
reproduzido fielmente (mesmo layout de colunas, bloco de metadados
repetido, linhas de Saldo Diário) a partir do que foi validado
manualmente contra o arquivo real durante o desenvolvimento.
"""

import io
from pathlib import Path

import openpyxl
import pytest

CLIENT_NAME = "Fulano De Tal"


class FakeUploadedFile:
    """Simula o objeto que o Streamlit file_uploader entrega: tem .name e .read()."""

    def __init__(self, data: bytes, name: str):
        self._data = data
        self.name = name

    def read(self) -> bytes:
        return self._data


def _write_metadata_block(ws, start_row: int, client_name: str, saldo_atual: float | None = None):
    """Bloco de Cliente/CPF/Agência/Conta/Período — igual ao que aparece
    (repetido) no extrato real do BTG em Excel."""
    ws.cell(row=start_row, column=2, value="Extrato de conta corrente")
    ws.cell(row=start_row + 2, column=2, value="Cliente:")
    ws.cell(row=start_row + 2, column=3, value=client_name)
    ws.cell(row=start_row + 3, column=2, value="CPF:")
    ws.cell(row=start_row + 3, column=3, value="000.000.000-00")
    ws.cell(row=start_row + 4, column=2, value="Agência:")
    ws.cell(row=start_row + 4, column=3, value="1")
    ws.cell(row=start_row + 5, column=2, value="Conta:")
    ws.cell(row=start_row + 5, column=3, value="12345-6")
    ws.cell(row=start_row + 6, column=2, value="Período do extrato:")
    ws.cell(row=start_row + 6, column=3, value="01/06/2026 a 30/06/2026")
    if saldo_atual is not None:
        ws.cell(row=start_row + 8, column=2, value="Lançamentos:")
        ws.cell(row=start_row + 8, column=8, value="Saldo atual:")
        ws.cell(row=start_row + 8, column=10, value=saldo_atual)
    return start_row + 10  # próxima linha livre


def _write_header_row(ws, row: int):
    ws.cell(row=row, column=2, value="Data e hora")
    ws.cell(row=row, column=3, value="Categoria")
    ws.cell(row=row, column=4, value="Transação")
    ws.cell(row=row, column=7, value="Descrição")
    ws.cell(row=row, column=11, value="Valor")


def build_synthetic_btg_xlsx(
    rows: list[dict],
    client_name: str = CLIENT_NAME,
    repeat_metadata_block: bool = True,
) -> bytes:
    """
    Monta um .xlsx no mesmo layout do extrato real do BTG Pactual (bloco
    de metadados + tabela de lançamentos, com colunas B=Data e hora,
    C=Categoria, D=Transação, G=Descrição, K=Valor).

    `rows`: lista de dicts com chaves datetime/categoria/transacao/descricao/valor.
    Uma linha de "Saldo Diário" tem categoria=None, transacao=None.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extrato"

    next_row = _write_metadata_block(ws, start_row=1, client_name=client_name, saldo_atual=rows[-1]["valor"] if rows else None)
    header_row = next_row
    _write_header_row(ws, header_row)

    r = header_row + 1
    for row in rows:
        ws.cell(row=r, column=2, value=row["datetime"])
        if row.get("categoria") is not None:
            ws.cell(row=r, column=3, value=row["categoria"])
        if row.get("transacao") is not None:
            ws.cell(row=r, column=4, value=row["transacao"])
        ws.cell(row=r, column=7, value=row["descricao"])
        ws.cell(row=r, column=11, value=row["valor"])
        r += 1

    if repeat_metadata_block:
        r += 2
        next_row2 = _write_metadata_block(ws, start_row=r, client_name=client_name)
        _write_header_row(ws, next_row2)
        r = next_row2 + 2
        ws.cell(row=r, column=6, value="Ouvidoria: 0800-000-0000")
        r += 1
        ws.cell(row=r, column=6, value="BTG Pactual - CNPJ 00.000.000/0000-00")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def client_name() -> str:
    return CLIENT_NAME


@pytest.fixture
def sample_btg_xlsx_bytes(client_name) -> bytes:
    """
    Extrato sintético representativo: 4 transações normais, 2
    transferências para o próprio titular (mesmo padrão do caso real que
    motivou a detecção de self_transfer), e 3 linhas de Saldo Diário
    intercaladas — 6 transações reais no total, 3 saldos.
    """
    rows = [
        {"datetime": "01/06/2026 07:35", "categoria": "Supermercado", "transacao": "Compra no débito autorizada", "descricao": "Amigao Brasil", "valor": -31.67},
        {"datetime": "01/06/2026 14:36", "categoria": "Transporte", "transacao": "Compra no débito autorizada", "descricao": "Uber", "valor": -7.54},
        {"datetime": "01/06/2026 22:18", "categoria": "Transferência", "transacao": "Pix enviado", "descricao": "Maria Jose Vieira", "valor": -220.0},
        {"datetime": "01/06/2026 23:59", "categoria": None, "transacao": None, "descricao": "Saldo Diário", "valor": 1223.40},
        {"datetime": "02/06/2026 09:00", "categoria": "Transferência", "transacao": "Transferência enviada", "descricao": client_name, "valor": -1000.0},
        {"datetime": "02/06/2026 10:00", "categoria": "Transferência", "transacao": "Transferência recebida", "descricao": client_name, "valor": 300.0},
        {"datetime": "02/06/2026 23:59", "categoria": None, "transacao": None, "descricao": "Saldo Diário", "valor": 464.19},
        {"datetime": "03/06/2026 11:11", "categoria": "Alimentação", "transacao": "Compra no débito autorizada", "descricao": "Restaurante Almoco", "valor": -21.79},
        {"datetime": "03/06/2026 23:59", "categoria": None, "transacao": None, "descricao": "Saldo Diário", "valor": 442.40},
    ]
    return build_synthetic_btg_xlsx(rows, client_name=client_name)


@pytest.fixture
def sample_btg_csv_bytes() -> bytes:
    content = (
        "Data;Descrição;Valor\n"
        "01/06/2026;PIX RECEBIDO JOAO SILVA;2500,00\n"
        "02/06/2026;SUPERMERCADO PAO DE ACUCAR;-345,67\n"
        "03/06/2026;UBER TRIP SAO PAULO;-28,50\n"
        "04/06/2026;NETFLIX.COM;-55,90\n"
    )
    return content.encode("utf-8-sig")


@pytest.fixture
def tmp_db_path(tmp_path: Path) -> Path:
    return tmp_path / "test_financas.db"
